import logging
import numpy as np
from django.views import View
from django.utils import timezone
from django.shortcuts import render, redirect
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin

from home.models import MyRogaCategory, LogModule, Logs
from my_roga.common.utils import read_file, VALID_FILE_EXTENSION
from my_roga.common.actions import Actions


logger = logging.getLogger(__name__)


class HomeView(View):
    template_name = "home/home.html"

    def get(self, request):
        return render(request, template_name=self.template_name, context={})


class UploadCategory(LoginRequiredMixin, View):
    template_name = "home/category/category.html"

    def get(self, request):
        records = MyRogaCategory.objects.isNotDeleted()
        return render(request, template_name=self.template_name, context={
            "error": False,
            "message": "Category(s) fetched Successfully!",
            "records": records
        })
    
    def post(self, request):
        VALID_COLUMNS = ['name', 'description', 'action', 'category id']
        upload_file = request.FILES.get("upload_file")
        file_extension = upload_file.name.split(".")[1]
        isAllColumnsValid = True

        # Check valid file extension
        if file_extension not in VALID_FILE_EXTENSION:
            logger.error("Please Upload file with only xls, xlsx & csv extension!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Please Upload file with only xls, xlsx & csv extension!"
            })
        
        # Read the uploaded file
        try:
            categories = read_file(upload_file)
        except Exception as e:
            logger.error(f"File not valid or does not have valid data! - {e.message}")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "File not valid or does not have valid data!"
            })
        
        # Uploaded file column names
        file_columns =  list(map(lambda x:x.lower(), list(categories.columns)))

        for column in file_columns:
            if column not in VALID_COLUMNS:
                isAllColumnsValid = False
                break

        if not isAllColumnsValid:
            logger.error("Uploaded file does not have valid column names!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Uploaded file does not have valid column names!"
            })

        # Replace nan to blank string
        categories.columns = file_columns
        categories = categories.replace(np.nan, '', regex=True)
        categories = categories.to_dict('records')

        # Create category object and store it into database
        for category in categories:
            request_params = {
                'name': category['name'],
                'description': category['description']
            }

            # Handle Actions
            if category['action'].upper() == 'CREATE':
                if MyRogaCategory.objects.filter(name__iexact=category['name']).exists():
                    logger.error(f"{category['name']} category already exist!")
                    continue
                Actions(MyRogaCategory).create(**request_params)
            elif category['action'].upper() == 'UPDATE':
                if category['category id']:
                    Actions(MyRogaCategory).update(category['category id'], **request_params)
                else:
                    logger.error(f"UPDATE - {category['name']} - Category ID {category['category id']} not exist!")
            elif category['action'].upper() == 'DELETE':
                if category['category id']:
                    Actions(MyRogaCategory).soft_delete(category['category id'])
                else:
                    logger.error(f"DELETE - {category['name']} - Category ID {category['category id']} not exist!")

        return redirect('/category-upload/')


class ExportCategoryToXlsx(LoginRequiredMixin, View):

    def get(self, request):
        """
        Downloads all categories as Excel file with a single worksheet
        """
        categories = MyRogaCategory.objects.isNotDeleted()
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-categories.xlsx'.format(
            date=timezone.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Categories'

        # Define the titles for columns
        columns = [
            'ID',
            'Name',
            'Description',
            'Updated At'
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all categories
        for category in categories:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                category.id,
                category.name,
                category.description,
                category.updated_at,
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class LogView(LoginRequiredMixin, View):
    template_name = "home/logs/logs.html"

    def get(self, request):

        selected_module = request.GET['module'] if 'module' in request.GET else None
        modules = LogModule.objects.isNotDeleted()
        if selected_module:
            records = Logs.objects.isNotDeleted().filter(log_module__id=selected_module)
            selected_module = modules.filter(id=selected_module).first()
        else:
            records = Logs.objects.isNotDeleted()
            selected_module = ""

        return render(request, template_name=self.template_name, context={
            "error": False,
            "message": "Category(s) fetched Successfully!",
            "records": records,
            "modules": modules,
        })