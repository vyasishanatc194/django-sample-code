import logging
import copy
import numpy as np
from django.utils import timezone
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views import View
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin

from rest_framework.views import APIView
from rest_framework.response import Response

from product.models import ProductCatalog, AspectsMetaData, AspectValuesMetaData, \
    AggregationModelSet
from model_catalog.models import ModelCatalog
from home.models import MyRogaCategory, Language, MarketPlace
from my_roga.common.utils import read_file, VALID_FILE_EXTENSION, MODULES_NAME
from my_roga.common.actions import Actions
from my_roga.common.logs import LogAction
from model_catalog.common import get_formatted_model_catalogs
from product.common import get_model_catalog_formatted_products


logger = logging.getLogger(__name__)


class ManageModelCatalog(LoginRequiredMixin, View):
    template_name = "model_catalog/model_catalog.html"

    def get(self, request):
        selected_category = request.GET['category'] if 'category' in request.GET else None
        categories = MyRogaCategory.objects.isNotDeleted()

        if selected_category:
            records = ModelCatalog.objects.isNotDeleted().filter(my_roga_category__id=selected_category)
            selected_category = categories.filter(id=selected_category).first()
        else:
            records = ModelCatalog.objects.isNotDeleted()
            selected_category = ""
        
        selected_category_id = selected_category.id if selected_category else ""
        response = get_formatted_model_catalogs(records, selected_category_id)
        results = response['results']
        columns = response['COLUMNS']
        return render(request, template_name=self.template_name, context={
            "error": False,
            "message": "Model Catalog fetched Successfully!",
            "records": results,
            "categories": categories,
            "columns": columns,
            "selected_category": selected_category,
        })

    def post(self, request):
        VALID_COLUMNS = ['action', 'model catalog id', 'model stealth', \
            'model status', 'model description', 'model description image url', \
            'model image url', 'recommendation score']
        upload_file = request.FILES.get("upload_file")
        file_extension = upload_file.name.split(".")[1]
        isAllColumnsValid = True

        # Check valid file extension
        if file_extension not in VALID_FILE_EXTENSION:
            logger.error("Please Upload file with only xls, xlsx & csv extension!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Please Upload file with only xls, xlsx & csv extension!"
            })
        
        # Read the uploaded file
        try:
            model_catalog = read_file(upload_file)
        except Exception as e:
            logger.error(f"File not valid or does not have valid data! - {e.message}")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "File not valid or does not have valid data!"
            })
        
        # Uploaded file column names
        file_columns =  list(map(lambda x:x.lower(), list(model_catalog.columns)))

        for column in file_columns:
            if column not in VALID_COLUMNS:
                isAllColumnsValid = False
                break
        
        if not isAllColumnsValid:
            logger.error("Uploaded file does not have valid column names!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Uploaded file does not have valid column names!"
            })
        
        model_catalog.columns = file_columns
        model_catalog = model_catalog.replace(np.nan, '', regex=True)
        model_catalog = model_catalog.to_dict('records')

        # Create Model Catalog object and store it into database
        for item in model_catalog:
            model_catalog_id = item['model catalog id']
            # Check If Model Catalog Id is valid or not
            if not model_catalog_id:
                logger.error(f"Model Catalog ID {model_catalog_id} is not valid or available in column!")
            
            model_catalog_action = item['action'].upper()

            # Get Model Catalog Object
            model_catalog_object = ModelCatalog.objects.filter(id=model_catalog_id).first()
            if not model_catalog_object:
                logger.error(f"{model_catalog_action} - Model Catalog with ID {model_catalog_id} is not available!")
            
            
            # Handle Model Catalog operations.
            if model_catalog_action == 'UPDATE':
                # Check Model Status value
                model_status = item['model status']
                if model_status.lower() == 'live':
                    model_status = '1'
                elif model_status.lower() == 'delete' or model_status.lower() == 'deleted':
                    model_status = '2'
                else:
                    model_status = ''
                # Request params for Update
                request_params = {
                    'model_stealth': item['model stealth'] if item['model stealth'] else '',
                    'model_status': model_status,
                    'description': item['model description'],
                    'description_image_url': item['model description image url'],
                    'image_url': item['model image url'],
                    'recommendation_score': item['recommendation score'] if item['recommendation score'] else 0,
                }
                Actions(ModelCatalog).update(model_catalog_id, **request_params)
            elif model_catalog_action == 'DELETE':
                delete_reason = 'Manual deletion'
                Actions(ModelCatalog).soft_delete(model_catalog_id, delete_reason)
            elif model_catalog_action == 'REVIVE':
                Actions(ModelCatalog).revive(model_catalog_id)
            else:
                logger.error(f"{model_catalog_action} Model Catalog Action not valid!")
        return redirect('/model-catalog/')


class ExportModelCatalog(LoginRequiredMixin, View):

    def get(self, request):
        """
        Downloads all Model Catalog as Excel file with a single worksheet
        """
        category_id = request.GET['category'] if 'category' in request.GET else None
        if category_id:
            model_catalog = ModelCatalog.objects.isNotDeleted().filter(my_roga_category=category_id)
        else:
            model_catalog = ModelCatalog.objects.isNotDeleted()
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-model_catalog.xlsx'.format(
            date=timezone.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Model Catalog'

        # Define the titles for columns
        columns = [
            'ID',
            'Model Title',
            'MyRogaCategoryID',
            'Aggregation Key',
            'Language',
            'Model Stealth',
            'Model Status',
            'Model Description',
            'Model Description Image URL',
            'Model Image URL',
            'Hero Product',
            'Hero Product Price',
            'Hero Product Market Place',
            'Second Hero Product',
            'Second Hero Product Price',
            'Second Hero Product Market Place',
            'Product Review Count',
            'Product Review Score',
            'Review URL',
            'Recommendation Score',
        ]
        row_num = 1

        # get formatted model catalog results
        response_results = get_formatted_model_catalogs(model_catalog, category_id)
        results = response_results['results']
        aspect_columns = response_results['COLUMNS']

        for column in aspect_columns:
            columns.append(column)

        columns.append('Updated At')

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all model catalog
        for model in results:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                model['id'],
                model['title'],
                model['my_roga_category'].id,
                model['aggregation_key'],
                model['language'],
                model['model_stealth'],
                model['model_status'],
                model['description'],
                model['description_image_url'],
                model['image_url'],
                model['hero_product'],
                model['hero_product_price'],
                model['hero_product_market_place'],
                model['second_hero_product'],
                model['second_hero_product_price'],
                model['second_hero_product_market_place'],
                model['product_review_count'],
                model['product_review_score'],
                model['review_url'],
                model['recommendation_score'],
            ]

            for col in aspect_columns:
                row.append(model[col])

            row.append(model['updated_at'])
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


# Model title formula - e.g. LG SX3232 55Inch Screen
def generate_model_title(model_catalog_item):
    model_catalog_items = copy.deepcopy(model_catalog_item)
    # remove products from item as it is not required for title
    model_catalog_items.pop('products')
    model_title_items = []
    model_title = ""
    # take Model Aspects from Aspect meta data of Aggregation key
    for aspect_value in model_catalog_items.values():
        model_title_item = {}
        is_model_aspect = aspect_value.aspect_meta_data.is_model_aspect

        # take model title items based on Aspect meta data only if it is model aspect
        if is_model_aspect:
            aspect_value_name = aspect_value.value_name if aspect_value.value_name else ''

            aspect_title_order = aspect_value.aspect_meta_data.model_title_order
            aspect_title_order = int(aspect_title_order) if aspect_title_order else 0

            text_before = aspect_value.aspect_meta_data.model_title_text_before_aspect
            text_before = text_before if text_before else ''

            text_after = aspect_value.aspect_meta_data.model_title_text_after_aspect
            text_after = text_after if text_after else ''

            model_title_item['text'] = aspect_value_name
            model_title_item['order'] = aspect_title_order
            model_title_item['text_before'] = text_before
            model_title_item['text_after'] = text_after
            model_title_items.append(model_title_item)

    # Sort model title items
    model_title_items = sorted(model_title_items, key = lambda item: item['order'])

    # Format model title according to title order and before/After text
    for title_item in model_title_items:
        title_text_before = f"{title_item['text_before']} " if title_item['text_before'] else ''
        title_text = f"{title_item['text']} " if title_item['text'] else ''
        title_text_after = f"{title_item['text_after']} " if title_item['text_after'] else ''
        model_title += f"{title_text_before}{title_text}{title_text_after}"
    return model_title.strip()


# Calculate Hero Product & Second Hero Product
# With Low price Diffrent Market place
def get_hero_product(products):
    products_query_set = copy.deepcopy(products)
    compare_products = {}
    for product in products_query_set:
        compare_products[product.id] = min(product.hero_offer_price, product.second_hero_offer_price)
    min_product_key = min(compare_products, key=compare_products.get)
    return products_query_set.filter(id=min_product_key).first()

def get_second_hero_product(hero_product, products):
    products_query_set = copy.deepcopy(products)

    # Check Offer from Hero Product which have lowe price and its market place.
    # We do not need to consider market place which is in Hero product offer with low price
    if hero_product.hero_offer_price < hero_product.second_hero_offer_price:
        hero_offer_market_place = hero_product.hero_offer_market_place
    else:
        hero_offer_market_place = hero_product.second_hero_offer_market_place
    if not hero_offer_market_place:
        logger.error(f"{hero_product.title} Hero product market place not available!")
        return None
    
    compare_products = {}
    for product in products_query_set:
        # We will have Offer from Different market place.
        # If Market place is same then assign minimum value as another market place
        if product.hero_offer_market_place == hero_offer_market_place:
            compare_products[product.id] = product.second_hero_offer_price
        elif product.second_hero_offer_market_place == hero_offer_market_place:
            compare_products[product.id] = product.hero_offer_price
        else:
            compare_products[product.id] = min(product.hero_offer_price, product.second_hero_offer_price)
    min_product_key = min(compare_products, key=compare_products.get)
    return products_query_set.filter(id=min_product_key).first()


# Calculate Model Catalog fields
def calculate_model_catalog_fields(model_catalog_item):
    # Create Model title
    model_title = generate_model_title(model_catalog_item)
    products = model_catalog_item['products']
    hero_product = get_hero_product(products)
    second_hero_product = get_second_hero_product(hero_product, products)
    
    # format and validate Model Catalog data
    hero_product_price = 0
    product_review_count = 0
    product_review_score = 0
    hero_product_market_place = None
    review_url = ''
    image_url = ''

    # Check and get valid Hero Product related data
    if hero_product:
        hero_product_price = hero_product.hero_offer_price
        hero_product_market_place = hero_product.hero_offer_market_place if hero_product.hero_offer_market_place else None
        review_url = hero_product.review_url if hero_product.review_url else ''
        image_url = hero_product.image_url if hero_product.image_url else ''
        product_review_count = hero_product.review_count if hero_product.review_count else 0
        product_review_score = hero_product.review_score if hero_product.review_score else 0

    # Check and get valid Second Hero Product related data
    second_hero_product_price = 0
    second_hero_product_market_place = None
    if hero_product and second_hero_product:
        if second_hero_product.hero_offer_market_place == hero_product_market_place:
            second_hero_product_price = second_hero_product.second_hero_offer_price
            second_hero_product_market_place = second_hero_product.second_hero_offer_market_place
        elif second_hero_product.second_hero_offer_market_place == hero_product_market_place:
            second_hero_product_price = second_hero_product.hero_offer_price
            second_hero_product_market_place = second_hero_product.hero_offer_market_place

    return {
        'model_title': model_title,
        'hero_product': hero_product,
        'hero_product_price': hero_product_price,
        'hero_product_market_place': hero_product_market_place,
        'second_hero_product': second_hero_product,
        'second_hero_product_price': second_hero_product_price,
        'second_hero_product_market_place': second_hero_product_market_place,
        'product_review_count': product_review_count,
        'product_review_score': product_review_score,
        'review_url': review_url,
        'image_url': image_url,
        'products': products,
    }


# Get Model Catalog Items based on Aggragation key
def get_model_catalog_items(aggregation_key):
    # Aggregation Aspects
    aggregation_aspects = aggregation_key.aspects.all()
    category = aggregation_key.my_roga_category
    
    # Model catalog language - Currently English and id is 1
    language =  Language.objects.filter(id=1).first()

    aggregation_products = []
    model_catalog_items = []

    # Get only those products which have same category as aggregation key
    products = ProductCatalog.objects.isNotDeleted().filter(my_roga_category=category)

    # Get Products which have aspects same as aggregation key
    for product in products:
        product_aspects = product.technical_aspects.all()

        # Check if Aggregation aspects is subset of Product technical aspects
        isSubset =  all(elem in product_aspects for elem in aggregation_aspects)
        if isSubset:
            model_catalog_data = {}

            # Create dict to get Aspect name and its value
            for aspect in aggregation_aspects:
                aspect_value = product.technical_aspect_values \
                    .filter(aspect_meta_data=aspect).first()
                if aspect_value:
                    model_catalog_data[aspect.name] = aspect_value if aspect_value else ''
            
            # Check if model aspect already available 
            isModelAspectAlreadyAvailable = False
            for model in model_catalog_items:
                if model == model_catalog_data:
                    isModelAspectAlreadyAvailable = True
            
            # if not model aspect already available then ad it into the model catalog products
            if not isModelAspectAlreadyAvailable:
                model_catalog_items.append(model_catalog_data)
            aggregation_products.append(product.id)

    # Create Filter to get model catalog products according to aspects and filter
    aggregation_products = products.filter(id__in=aggregation_products)
    for model in model_catalog_items:
        query_filter = aggregation_products

        # get all valid products based on aggragation keys
        isAllValidProductAspect = True
        for key, value in model.items():
            filtered_products = query_filter.filter(technical_aspect_values=value)
            query_filter = filtered_products
            if not filtered_products:
                isAllValidProductAspect = False
        
        # Add products to filter if all aggragation aspect valid with product aspects and values
        model['products'] = []
        if isAllValidProductAspect:
            model['products'] = query_filter

    return {
        'aggregation_key': aggregation_key,
        'category': category,
        'language': language,
        'model_catalog_items': model_catalog_items,
    }


# Create Model Catalog based on the Aggregation key
def create_model_catalog(aggregation_key):
    # Get Model Catalog items which needs to store inside of Model
    model_catalog_items = get_model_catalog_items(aggregation_key)

    # Get Calcualted Model Catalog fields
    for model_catalog_item in model_catalog_items['model_catalog_items']:
        model_catalog_fields = calculate_model_catalog_fields(model_catalog_item)

        model_aggregation_key = "; ".join([aspect.name for aspect in model_catalog_items['aggregation_key'].aspects.all()])
        # Model Catalog params
        model_catalog_params = {
            'aggregation_model_set': model_catalog_items['aggregation_key'],
            'model_aggregation_key': model_aggregation_key,
            'my_roga_category': model_catalog_items['category'],
            'language': model_catalog_items['language'],
            'title': model_catalog_fields['model_title'],
            'hero_product': model_catalog_fields['hero_product'],
            'hero_product_price': model_catalog_fields['hero_product_price'],
            'hero_product_market_place': model_catalog_fields['hero_product_market_place'],
            'second_hero_product': model_catalog_fields['second_hero_product'],
            'second_hero_product_price': model_catalog_fields['second_hero_product_price'],
            'second_hero_product_market_place': model_catalog_fields['second_hero_product_market_place'],
            'product_review_count': model_catalog_fields['product_review_count'],
            'product_review_score': model_catalog_fields['product_review_score'],
            'review_url': model_catalog_fields['review_url'],
            'image_url': model_catalog_fields['image_url'],
        }

        # Create Model Catalog
        model_catalog = Actions(ModelCatalog).create(**model_catalog_params)

        # Map Product Catalog object to Model Catalog
        for product in model_catalog_fields['products']:
            model_catalog.product_catalog.add(product)
        model_catalog.save()
        
        
class FetchModelCatalogProducts(APIView):
    def get(self, request, id):
        model_catalog = ModelCatalog.objects.filter(id=id).first()
        response_output = {
            'columns': [],
            'results': [],
        }
        results = []
        if model_catalog:
            products = model_catalog.product_catalog.all()
            category = model_catalog.my_roga_category
            response = get_model_catalog_formatted_products(products, category)
            response_output['columns'] = response['COLUMNS']
            response_output['results'] = response['results']
        return Response(response_output)


# Find If any change in existing Model catalog values
def is_model_catalog_change():
    pass


# Reprocess Model Catalog
class ReprocessModelCatalog(APIView):
    def get(self, request, category_id):
        response_output = {}
        # Get Aggregation key by category id
        aggregation_key_obj = AggregationModelSet.objects.filter(my_roga_category__id=category_id).first()
        if aggregation_key_obj:
            # Get All the model catalog by category filter and Aggregation model set
            model_catalogs = ModelCatalog.objects.filter(my_roga_category__id=category_id, aggregation_model_set=aggregation_key_obj)
            # Total models in this selected category
            total_models = model_catalogs
            # Live models in this selected category - Not deleted models
            live_models = model_catalogs.isNotDeleted()
            # Get Model Catalog item details from aggregation key
            model_catalog_items = get_model_catalog_items(aggregation_key_obj)
            print(model_catalog_items)
        else:
            logger.error(f"Aggregation Key | Aggregation key with Category ID {category_id} not available!")
        return Response(response_output)


"""
    Auto Update Model Catalog according to the Products Change.
    We will update following fields of Model Catalog according to Product Change.
        - product_catalog
        - hero_product
        - hero_product_price
        - hero_product_market_place
        - second_hero_product
        - second_hero_product_price
        - second_hero_product_market_place
        - product_review_count
        - product_review_score
        - review_url
        - updated_at
"""
def update_model_catalog():
    # Get All unique distinct Aggregation key from existing model
    aggregation_keys = ModelCatalog.objects.isNotDeleted().values_list('aggregation_model_set', \
        flat=True).distinct().order_by('aggregation_model_set')

    # Take each aggregation model set object from aggregation_key id
    for key in aggregation_keys:
        aggregation_key_obj = AggregationModelSet.objects.filter(id=key).first()
        if not aggregation_key_obj:
            logger.error(f"Update Model Catalog After Product Upload | Aggregation key with ID {key} not available!")

        # Get Model Catalog item details
        model_catalog_items = get_model_catalog_items(aggregation_key_obj)

        aggregation_model_set = model_catalog_items['aggregation_key']
        my_roga_category = model_catalog_items['category']

        # Get Calcualted Model Catalog fields
        for model_catalog_item in model_catalog_items['model_catalog_items']:
            model_catalog_fields = calculate_model_catalog_fields(model_catalog_item)
            model_catalog_products = model_catalog_fields['products']

            # Here Every Model will have unique products as Product will not repeat in same category and aggregation key model
            model_catalog_obj = ModelCatalog.objects.isNotDeleted().filter(aggregation_model_set=aggregation_model_set, \
                my_roga_category=my_roga_category, product_catalog__in=model_catalog_products).first()

            if model_catalog_obj:
                # Model Catalog params which needs to update
                model_catalog_params = {
                    'hero_product': model_catalog_fields['hero_product'],
                    'hero_product_price': model_catalog_fields['hero_product_price'],
                    'hero_product_market_place': model_catalog_fields['hero_product_market_place'],
                    'second_hero_product': model_catalog_fields['second_hero_product'],
                    'second_hero_product_price': model_catalog_fields['second_hero_product_price'],
                    'second_hero_product_market_place': model_catalog_fields['second_hero_product_market_place'],
                    'product_review_count': model_catalog_fields['product_review_count'],
                    'product_review_score': model_catalog_fields['product_review_score'],
                    'review_url': model_catalog_fields['review_url'],
                }

                # Check Change Model Catalog values
                isSameModelCatalogValues = False
                if model_catalog_obj.hero_product == model_catalog_params['hero_product'] and \
                    model_catalog_obj.hero_product_price == model_catalog_params['hero_product_price'] and \
                    model_catalog_obj.hero_product_market_place == model_catalog_params['hero_product_market_place'] and \
                    model_catalog_obj.second_hero_product == model_catalog_params['second_hero_product'] and \
                    model_catalog_obj.second_hero_product_price == model_catalog_params['second_hero_product_price'] and \
                    model_catalog_obj.second_hero_product_market_place == model_catalog_params['second_hero_product_market_place'] and \
                    model_catalog_obj.product_review_count == model_catalog_params['product_review_count'] and \
                    model_catalog_obj.product_review_score == model_catalog_params['product_review_score'] and \
                    model_catalog_obj.review_url == model_catalog_params['review_url']:
                    isSameModelCatalogValues = True
                else:
                    isSameModelCatalogValues = False

                model_catalog_id = model_catalog_obj.id
                if not isSameModelCatalogValues:
                    # Update Model Catalog
                    updated_model_catalog = Actions(ModelCatalog).update(model_catalog_id, **model_catalog_params)

                    # Map Updated Product Catalog object to Model Catalog
                    for product in model_catalog_fields['products']:
                        updated_model_catalog.product_catalog.add(product)
                    updated_model_catalog.save()
                    logger.info(f"Update Model Catalog | Model Catalog ID {model_catalog_id} Updated successfully!")
                else:
                    logger.info(f"Update Model Catalog | Model Catalog ID {model_catalog_id} has no change!")
            else:
                logger.error(f"Update Model Catalog | Model Catalog not available for {model_catalog_item['model_title']} !")
    return False