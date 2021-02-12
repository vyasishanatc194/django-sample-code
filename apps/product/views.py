import logging
import copy
import numpy as np
from django.utils import timezone
from django.shortcuts import render, redirect
from django.views import View
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin

from product.models import ProductCatalog, AspectsMetaData, AspectValuesMetaData, \
    AggregationModelSet
from home.models import MyRogaCategory, Language, MarketPlace
from my_roga.common.utils import read_file, VALID_FILE_EXTENSION, MODULES_NAME
from my_roga.common.actions import Actions
from my_roga.common.logs import LogAction
from product.common import get_formatted_products
from model_catalog.views import create_model_catalog, update_model_catalog


logger = logging.getLogger(__name__)


class UploadProducts(LoginRequiredMixin, View):
    template_name = "product/product.html"

    def get(self, request):
        selected_category = request.GET['category'] if 'category' in request.GET else None
        selected_product = request.GET['product'] if 'product' in request.GET else None

        categories = MyRogaCategory.objects.isNotDeleted()
        products = ProductCatalog.objects.isNotDeleted()

        # Get Product Columns dynamically and its value
        response = get_formatted_products(categories, products, selected_category, selected_product)

        COLUMNS = response['COLUMNS']
        results = response['results']
        selected_product = response["selected_product"]
        selected_category = response["selected_category"]
        return render(request, template_name=self.template_name, context={
            "error": False,
            "message": "Product(s) fetched Successfully!",
            "records": results,
            "categories": categories,
            "products": products,
            "selected_category": selected_category,
            "selected_product": selected_product,
            "categories": categories,
            "columns": COLUMNS,
        })

    def add_technical_aspect(self, **kwargs):
        TECHNICAL_ASPECTS_VALID_COLUMNS = kwargs['TECHNICAL_ASPECTS_VALID_COLUMNS']
        product = kwargs['product']
        product_catalog = kwargs['product_catalog']
        category = kwargs['category']
        for key, value in product.items():
            # Check if field is available in technical aspect
            if key in TECHNICAL_ASPECTS_VALID_COLUMNS:
                technical_aspect = AspectsMetaData.objects.filter(name__iexact=key, my_roga_category=category).first()
                product_catalog.technical_aspects.add(technical_aspect)

                if AspectValuesMetaData.objects.filter(
                        aspect_meta_data=technical_aspect,
                        my_roga_category=category,
                        value_name=product[key]).exists():

                    technical_aspect_value = AspectValuesMetaData.objects.filter(
                                                aspect_meta_data=technical_aspect,
                                                my_roga_category=category,
                                                value_name=product[key]).first()
                else:
                    technical_aspect_value = AspectValuesMetaData.objects.create(
                        aspect_meta_data = technical_aspect,
                        my_roga_category = category,
                        value_name = product[key],
                        value_type = '1' # 1 - Text
                    ) # TODO value_type needs to deside - Currently set it to TEXT
                product_catalog.technical_aspect_values.add(technical_aspect_value)
        product_catalog.save()
        return product_catalog
    
    def check_valid_product_technical_aspect(self, **kwargs):
        isValidTechnicalAspect = True
        TECHNICAL_ASPECTS_VALID_COLUMNS = kwargs['TECHNICAL_ASPECTS_VALID_COLUMNS']
        product = kwargs['product']
        category = kwargs['category']
        notValidTechnicalAspects = []
        for key, value in product.items():
            if key in TECHNICAL_ASPECTS_VALID_COLUMNS:
                technical_aspect = AspectsMetaData.objects.filter(name__iexact=key, my_roga_category=category).first()
                if not technical_aspect:
                    isValidTechnicalAspect = False
                    notValidTechnicalAspects.append(key)
        
        if not isValidTechnicalAspect:
            logger.error(f"{','.join(notValidTechnicalAspects)} - Technical Aspect with Category {category.id} not exist!")
            LogAction(MODULES_NAME['PRODUCT']).create('ERROR',
                f"{','.join(notValidTechnicalAspects)} - Technical Aspect with Category {category.id} not exist!")
            return False

        return True
    
    def post(self, request):
        VALID_COLUMNS = ['category', 'language', 'hero offer', 'second offer', \
            'title', 'description', 'product status', 'product stealth', \
            'action', 'product id', 'image url', 'review url', 'reviews count', 'reviews score', \
            'hero offer url', 'hero offer price', 'hero offer marketplace id', \
            'second offer url', 'second offer price', 'second offer marketplace id']

        PRODUCT_TECHNICAL_ASPECTS = list(AspectsMetaData.objects.values_list('name', flat=True))
        TECHNICAL_ASPECTS_VALID_COLUMNS = [ aspect.lower() for aspect in PRODUCT_TECHNICAL_ASPECTS]

        VALID_COLIMNS =  VALID_COLUMNS + TECHNICAL_ASPECTS_VALID_COLUMNS

        upload_file = request.FILES.get("upload_file")
        file_extension = upload_file.name.split(".")[1]
        isAllColumnsValid = True

        # Check valid file extension
        if file_extension not in VALID_FILE_EXTENSION:
            logger.error("Please Upload file with only xls, xlsx & csv extension!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Please Upload file with only xls, xlsx & csv extension!"
            })

        # Read the uploaded file
        try:
            products = read_file(upload_file)
        except Exception as e:
            logger.error(f"File not valid or does not have valid data! - {e.message}")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "File not valid or does not have valid data!"
            })
        
        # Uploaded file column names & Verify if it is valid
        file_columns =  list(map(lambda x:x.lower(), list(products.columns)))

        isAllColumnsValid = True
        for col in file_columns:
            if col not in VALID_COLIMNS:
                isAllColumnsValid = False
                break
        
        if not isAllColumnsValid:
            logger.error("Uploaded file does not have valid column names!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Uploaded file does not have valid column names!"
            })
        
        products.columns = file_columns
        # Replace nan to blank string
        products = products.replace(np.nan, '', regex=True)
        products = products.to_dict('records')

        for product in products:
            if not MyRogaCategory.objects.filter(id=product['category']).exists():
                logger.error(f"{product['title']} - Category ID {product['category']} not exist!")
                continue
            
            # Get Category of Product
            category = MyRogaCategory.objects.get(id=product['category'])

            # Check Technical aspect is Valid or not
            aspect_params = {
                    'product': product,
                    'TECHNICAL_ASPECTS_VALID_COLUMNS': TECHNICAL_ASPECTS_VALID_COLUMNS,
                    'category': category
                }

            isAspectsValid = self.check_valid_product_technical_aspect(**aspect_params)
            if not isAspectsValid:
                logger.error(f"{product['title']} - Technical Aspects not valid!")
                continue
            
            # Get Language of Product
            try:
                language = Language.objects.filter(name__iexact=product['language']).first()
            except:
                logger.error("Language not available or not have a column in uploaded file!")
                language = None

            hero_offer_market_place = None
            if 'hero offer marketplace id' in product.keys() and product['hero offer marketplace id']:
                hero_offer_market_place = MarketPlace.objects.filter(id=int(product['hero offer marketplace id'])).first()

            second_offer_market_place = None
            if 'second offer marketplace id' in product.keys() and product['second offer marketplace id']:
                second_offer_market_place = MarketPlace.objects.filter(id=int(product['second offer marketplace id'])).first()

            request_params = {
                'title': product['title'] if 'title' in product.keys() else '',
                'description': product['description'] if 'description' in product.keys() else '',
                'my_roga_category': category,
                'language': language,
                'hero_offer': None,
                'second_offer': None,
                'image_url': product['image url'] if 'image url' in product.keys() else '',
                'review_url': product['review url'] if 'review url' in product.keys() else '',
                'review_count': product['reviews count'] if 'reviews count' in product.keys() and product['reviews count'] else 0,
                'review_score': product['reviews score'] if 'reviews score' in product.keys() and product['reviews score'] else 0,
                'hero_offer_url': product['hero offer url'] if 'hero offer url' in product.keys() and product['hero offer url'] else '',
                'hero_offer_price': product['hero offer price'] if 'hero offer price' in product.keys() and product['hero offer price'] else 0,
                'hero_offer_market_place': hero_offer_market_place,
                'second_hero_offer_url': product['second offer url'] if 'second offer url' in product.keys() and product['second offer url'] else '',
                'second_hero_offer_price': product['second offer price'] if 'second offer price' in product.keys() and product['second offer price'] else 0,
                'second_hero_offer_market_place': second_offer_market_place,
                'product_status': '1' if product['product status'].lower() == 'live' else '2',
                'product_stealth': product['product stealth'] if 'product stealth' in product.keys() and product['product stealth'] else ''
            }

            product_id = product['product id'] if 'product id' in product.keys() else None

            if product['action'].upper() == 'CREATE':
                if not ProductCatalog.objects.filter(title__iexact=product['title'], description__iexact=product['description']).exists():
                    product_catalog = Actions(ProductCatalog).create(**request_params)
                    args = {
                        'product': product,
                        'TECHNICAL_ASPECTS_VALID_COLUMNS': TECHNICAL_ASPECTS_VALID_COLUMNS,
                        'product_catalog': product_catalog,
                        'category': category
                    }
                    # Set technical Aspect Values
                    self.add_technical_aspect(**args)
                else:
                    logger.error(f"{product['title']} - {product['product id']} product already exist!")

            elif product['action'].upper() == 'UPDATE':
                if product_id:
                    product_catalog = Actions(ProductCatalog).update(product_id, **request_params)
                    if product_catalog:
                        product_catalog.technical_aspect_values.clear()
                        args = {
                            'product': product,
                            'TECHNICAL_ASPECTS_VALID_COLUMNS': TECHNICAL_ASPECTS_VALID_COLUMNS,
                            'product_catalog': product_catalog,
                            'category': category
                        }
                        # Set technical Aspect Values
                        self.add_technical_aspect(**args)
                    else:
                        logger.error(f"UPDATE - {product['title']} - Product {product_id} not exist!")
                else:
                    logger.error(f"UPDATE - {product['title']} - Product ID {product_id} not exist!")
            elif product['action'].upper() == 'DELETE':
                if product_id:
                    product_catalog = ProductCatalog.objects.filter(id=int(product_id)).first()
                    if product_catalog:
                        # Soft Delete Existing Technical Aspect values
                        technical_aspect_values = product_catalog.technical_aspect_values.all()
                        for aspect_value in technical_aspect_values:
                            Actions(AspectValuesMetaData).soft_delete(aspect_value.id)
                    else:
                        logger.error(f"DELETE - {product['title']} - Product {product_id} not exist!")
                    Actions(ProductCatalog).soft_delete(product_id)
                else:
                    logger.error(f"DELETE - {product['title']} - Product ID {product_id} not exist!")
            
        # Update Model Catalog after Product catalog change.
        is_updated = update_model_catalog()
        if not is_updated:
            logger.info("No change in Model catalog after Products Upload Actions!")

        return redirect('/product/upload/')


class ExportProductsToXlsx(LoginRequiredMixin, View):

    def get(self, request):
        """
        Downloads all products as Excel file with a single worksheet
        """
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-products.xlsx'.format(
            date=timezone.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Products'

        selected_category = request.GET['category'] if 'category' in request.GET else None
        selected_product = request.GET['product'] if 'product' in request.GET else None
        
        categories = MyRogaCategory.objects.isNotDeleted()
        products = ProductCatalog.objects.isNotDeleted()

        # Get Product Columns dynamically and its value
        formatted_product_response = get_formatted_products(categories, products, selected_category, selected_product)

        COLUMNS = formatted_product_response['COLUMNS']
        results = formatted_product_response['results']

        # Define the titles for columns
        columns = COLUMNS
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Assign result object to Product
        products = results
        # Iterate through all products
        for product in products:
            row_num += 1

            # Define the data for each cell in the row 
            row = []
            for col in columns:
                row.append(product[col])

            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class UploadProductAspects(LoginRequiredMixin, View):
    template_name = 'product/aspects.html'

    def get(self, request):
        selected_category = request.GET['category'] if 'category' in request.GET else None
        categories = MyRogaCategory.objects.isNotDeleted()
        if selected_category:
            records = AspectsMetaData.objects.isNotDeleted().filter(my_roga_category=selected_category)
            selected_category = categories.filter(id=selected_category).first()
        else:
            records = AspectsMetaData.objects.isNotDeleted()
            selected_category = ""

        results = []
        for record in records:
            result = {}
            result['id'] = record.id
            result['name'] = record.name
            result['my_roga_category'] = record.my_roga_category
            result['aspect_type'] = record.get_aspect_type_display()
            result['aspect_input_type'] = record.get_aspect_input_type_display()
            result['is_model_aspect'] = 'Y' if record.is_model_aspect else ''
            result['model_title_order'] = record.model_title_order if record.model_title_order else ''
            result['model_title_text_before_aspect'] = record.model_title_text_before_aspect
            result['model_title_text_after_aspect'] = record.model_title_text_after_aspect
            result['updated_at'] = record.updated_at
            results.append(result)

        return render(request, template_name=self.template_name, context={
            "error": False,
            "message": "Aspect(s) & Category(s) fetched Successfully!",
            "records": results,
            "categories": categories,
            "selected_category": selected_category,
        })
    
    def post(self, request):
        VALID_COLUMNS = ['myrogacategoryid', 'aspectname', 'closed/open', \
            'multi/single', 'action', 'aspect id', 'is model aspect', \
            'model title order', 'model title text before aspect', \
            'model title text after aspect']
        upload_file = request.FILES.get("upload_file")
        file_extension = upload_file.name.split(".")[1]
        isAllColumnsValid = True

        # Check valid file extension
        if file_extension not in VALID_FILE_EXTENSION:
            logger.error("Please Upload file with only xls, xlsx & csv extension!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Please Upload file with only xls, xlsx & csv extension!"
            })
        
        # Read the uploaded file
        try:
            product_aspects = read_file(upload_file)
        except Exception as e:
            logger.error(f"File not valid or does not have valid data! - {e.message}")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "File not valid or does not have valid data!"
            })
        
        # Uploaded file column names
        file_columns =  list(map(lambda x:x.lower(), list(product_aspects.columns)))

        for column in file_columns:
            if column not in VALID_COLUMNS:
                isAllColumnsValid = False
                break
        
        if not isAllColumnsValid:
            logger.error("Uploaded file does not have valid column names!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Uploaded file does not have valid column names!"
            })

        product_aspects.columns = file_columns
        product_aspects = product_aspects.replace(np.nan, '', regex=True)
        product_aspects = product_aspects.to_dict('records')

        # Create product aspect object and store it into database
        for aspect in product_aspects:
            my_roga_category = MyRogaCategory.objects.filter(id=aspect['myrogacategoryid']).first()
            if not my_roga_category:
                logger.error(f"Category ID {aspect['myrogacategoryid']} not exist!")
                continue
            
            is_model_aspect = True if aspect['is model aspect'] == 'Y' else False

            model_title_order = aspect['model title order']
            model_title_order = int(model_title_order) if model_title_order else None

            model_title_text_before_aspect = aspect['model title text before aspect']
            model_title_text_before_aspect = model_title_text_before_aspect if model_title_text_before_aspect else ''

            model_title_text_after_aspect = aspect['model title text after aspect']
            model_title_text_after_aspect = model_title_text_after_aspect if model_title_text_after_aspect else ''


            # Request Params
            request_params = {
                'name': aspect['aspectname'],
                'my_roga_category': my_roga_category,
                'aspect_type': aspect['closed/open'],
                'aspect_input_type': aspect['multi/single'],
                'is_model_aspect': is_model_aspect,
                'model_title_order': model_title_order,
                'model_title_text_before_aspect': model_title_text_before_aspect,
                'model_title_text_after_aspect': model_title_text_after_aspect,
            }

            # Handle Actions
            if aspect['action'].upper() == 'CREATE':
                if AspectsMetaData.objects.filter(name__iexact=aspect['aspectname'],
                    my_roga_category=aspect['myrogacategoryid']).exists():
                    logger.error(f"{aspect['aspectname']} - Category ID {aspect['myrogacategoryid']} already exist!")
                    continue
                Actions(AspectsMetaData).create(**request_params)
            elif aspect['action'].upper() == 'UPDATE':
                if aspect['aspect id']:
                    Actions(AspectsMetaData).update(aspect['aspect id'], **request_params)
                else:
                    logger.error(f"UPDATE - {aspect['aspectname']} - Aspect ID {aspect['aspect id']} not exist!")
            elif aspect['action'].upper() == 'DELETE':
                if aspect['aspect id']:
                    Actions(AspectsMetaData).soft_delete(aspect['aspect id'])
                else:
                    logger.error(f"DELETE - {aspect['aspectname']} - Aspect ID {aspect['aspect id']} not exist!")

        return redirect('/product/aspect-upload/')


class ExportProductAspectToXlsx(LoginRequiredMixin, View):

    def get(self, request):
        """
        Downloads all Product Aspects as Excel file with a single worksheet
        """
        category_id = request.GET['category_id'] if 'category_id' in request.GET else None
        if category_id:
            product_aspects = AspectsMetaData.objects.isNotDeleted().filter(my_roga_category=category_id)
        else:
            product_aspects = AspectsMetaData.objects.isNotDeleted()
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-product_aspects.xlsx'.format(
            date=timezone.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Product Aspects'

        # Define the titles for columns
        columns = [
            'ID',
            'MyRogaCategoryID',
            'AspectName',
            'Closed/Open',
            'Multi/Single value',
            'Is Model Aspect',
            'Model Title Order',
            'Model Title Text Before Aspect',
            'Model Title Text After Aspect',
            'Updated At',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all product aspects
        for aspect in product_aspects:
            row_num += 1
            
            # Define the data for each cell in the row
            is_model_aspect = 'Y' if aspect.is_model_aspect else ''
            model_title_order = aspect.model_title_order if aspect.model_title_order else ''
            model_title_text_before_aspect = aspect.model_title_text_before_aspect
            model_title_text_after_aspect = aspect.model_title_text_after_aspect

            row = [
                aspect.id,
                aspect.my_roga_category.id,
                aspect.name,
                aspect.get_aspect_type_display(),
                aspect.get_aspect_input_type_display(),
                is_model_aspect,
                model_title_order,
                model_title_text_before_aspect,
                model_title_text_after_aspect,
                aspect.updated_at
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class UploadAggregationModelSet(LoginRequiredMixin, View):
    template_name = "product/aggregation_model_set.html"

    def get(self, request):
        selected_category = request.GET['category'] if 'category' in request.GET else None
        categories = MyRogaCategory.objects.isNotDeleted()
        if selected_category:
            records = AggregationModelSet.objects.isNotDeleted().filter(my_roga_category__id=selected_category)
            selected_category = categories.filter(id=selected_category).first()
        else:
            records = AggregationModelSet.objects.isNotDeleted()
            selected_category = ""
        
        results = []
        for record in records:
            result = {}
            result['id'] = record.id
            result['my_roga_category'] = record.my_roga_category
            result['aspects'] = "; ".join([aspect.name for aspect in record.aspects.all()])
            result['updated_at'] = record.updated_at
            results.append(result)

        return render(request, template_name=self.template_name, context={
            "error": False,
            "message": "Aggregation Model Set(s) fetched Successfully!",
            "records": results,
            "categories": categories,
            "selected_category": selected_category,
        })

    def post(self, request):
        VALID_COLUMNS = ['myrogacategoryid', 'aspects', 'action', 'aggregation id']
        upload_file = request.FILES.get("upload_file")
        file_extension = upload_file.name.split(".")[1]
        isAllColumnsValid = True

        # Check valid file extension
        if file_extension not in VALID_FILE_EXTENSION:
            logger.error("Please Upload file with only xls, xlsx & csv extension!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Please Upload file with only xls, xlsx & csv extension!"
            })
        
        # Read the uploaded file
        try:
            aggregation_model_set = read_file(upload_file)
        except Exception as e:
            logger.error(f"File not valid or does not have valid data! - {e.message}")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "File not valid or does not have valid data!"
            })
        
        # Uploaded file column names
        file_columns =  list(map(lambda x:x.lower(), list(aggregation_model_set.columns)))

        for column in file_columns:
            if column not in VALID_COLUMNS:
                isAllColumnsValid = False
                break
        
        if not isAllColumnsValid:
            logger.error("Uploaded file does not have valid column names!")
            return render(request, template_name=self.template_name, context={
                "error": True,
                "message": "Uploaded file does not have valid column names!"
            })

        aggregation_model_set.columns = file_columns
        aggregation_model_set = aggregation_model_set.replace(np.nan, '', regex=True)
        aggregation_model_set = aggregation_model_set.to_dict('records')

        # Create Aggregation model set object and store it into database
        for aggregation in aggregation_model_set:
            isAllAspectsValid = True

            my_roga_category = MyRogaCategory.objects.filter(id=aggregation['myrogacategoryid']).first()
            if not my_roga_category:
                logger.error(f"Aggregation Key Category ID {aggregation['myrogacategoryid']} not exist!")
                LogAction(MODULES_NAME['AGGREGATION_MODEL_SET']).create('ERROR',
                    f"Aggregation Key Category ID {aggregation['myrogacategoryid']} not exist!")
                continue
            
            aspects = [x.strip() for x in aggregation['aspects'].split(';')] if 'aspects' in aggregation.keys() else ''

            # Filter to check aspect name and category
            technical_aspects = []
            for aspect in aspects:
                technical_aspect_obj = AspectsMetaData.objects.filter(name__iexact=aspect, my_roga_category=my_roga_category).first()
                if not technical_aspect_obj:
                    isAllAspectsValid = False
                    technical_aspects = []
                    break
                technical_aspects.append(technical_aspect_obj)

            # Check if Aspect with Category Exist or not
            if not isAllAspectsValid:
                logger.error(f"Aggregation Key Aspect {';'.join(aspects)} with Category ID {aggregation['myrogacategoryid']} not exist!")
                LogAction(MODULES_NAME['AGGREGATION_MODEL_SET']).create('ERROR',
                    f"Aggregation Key Aspect {';'.join(aspects)} with Category ID {aggregation['myrogacategoryid']} not exist!")
                continue

            # Request Params
            request_params = {
                'my_roga_category': my_roga_category,
            }

            aggregation_id = aggregation['aggregation id'] if 'aggregation id' in aggregation.keys() else None

            # Handle Actions
            if aggregation['action'].upper() == 'CREATE':
                if AggregationModelSet.objects.filter(aspects__in=technical_aspects,
                    my_roga_category=aggregation['myrogacategoryid']).exists():
                    logger.error(f"Aggregation key {aggregation['aspects']} with Category ID {aggregation['myrogacategoryid']} already exist!")
                    LogAction(MODULES_NAME['AGGREGATION_MODEL_SET']).create('ERROR',
                        f"Aggregation key {aggregation['aspects']} with Category ID {aggregation['myrogacategoryid']} already exist!")
                    continue
                
                aggregation_model_set_obj = AggregationModelSet.objects.filter(my_roga_category=aggregation['myrogacategoryid']).first()
                if not aggregation_model_set_obj:
                    aggregation_model_set_obj = Actions(AggregationModelSet).create(**request_params)
                for aspect in technical_aspects:
                    aggregation_model_set_obj.aspects.add(aspect)
                aggregation_model_set_obj.save()

                # Map Aggregation key to Product Catalog and Create Model from that.
                create_model_catalog(aggregation_model_set_obj)
            elif aggregation['action'].upper() == 'UPDATE':
                if aggregation_id:
                    aggregation_model_set_obj = Actions(AggregationModelSet).update(aggregation_id, **request_params)
                    aggregation_model_set_obj.aspects.clear()
                    for aspect in technical_aspects:
                        aggregation_model_set_obj.aspects.add(aspect)
                    aggregation_model_set_obj.save()
                else:
                    logger.error(f"UPDATE - {aggregation['aspects']} - Aggregation ID {aggregation_id} not exist!")
            elif aggregation['action'].upper() == 'DELETE':
                if aggregation_id:
                    Actions(AggregationModelSet).soft_delete(aggregation_id)
                else:
                    logger.error(f"DELETE - {aggregation['aspects']} - Aggregation ID {aggregation_id} not exist!")

        return redirect('/product/aggregation-model-set/')


class ExportAggregationModelSet(LoginRequiredMixin, View):

    def get(self, request):
        """
        Downloads all Aggregation Model set as Excel file with a single worksheet
        """
        category_id = request.GET['category_id'] if 'category_id' in request.GET else None
        if category_id:
            aggregation_model_set = AggregationModelSet.objects.isNotDeleted().filter(my_roga_category=category_id)
        else:
            aggregation_model_set = AggregationModelSet.objects.isNotDeleted()
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-aggregation_model_set.xlsx'.format(
            date=timezone.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Aggregation Model Set'

        # Define the titles for columns
        columns = [
            'ID',
            'MyRogaCategoryID',
            'Aspects',
            'Updated At'
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all aggregation model set
        for aggregation in aggregation_model_set:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                aggregation.id,
                aggregation.my_roga_category.id,
                "; ".join([aspect.name for aspect in aggregation.aspects.all()]),
                aggregation.updated_at
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response

